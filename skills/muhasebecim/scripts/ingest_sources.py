#!/usr/bin/env python3
"""Local, content-addressed source ingestion for Muhasebecim.

The script may download public HTTPS documents. It never uploads local content.
Customer/case documents (scope=case) must be local filesystem paths.
"""

from __future__ import annotations

import argparse
import hashlib
import io
import json
import mimetypes
import re
import shutil
import subprocess
import sys
import tempfile
import urllib.error
import urllib.parse
import urllib.request
from collections import Counter
from datetime import date, datetime, timezone
from html.parser import HTMLParser
from pathlib import Path
from typing import Any


VERSION = "0.0.1"
STATUSES = {"in_force", "future", "draft", "repealed", "administrative_view", "unknown"}
SCOPES = {"public", "case"}
TEXT_TYPES = {".txt", ".md", ".csv", ".tsv", ".json", ".xml", ".xhtml"}


class IngestError(ValueError):
    pass


class VisibleTextParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.parts: list[str] = []
        self.suppressed = 0

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() in {"script", "style", "noscript", "svg"}:
            self.suppressed += 1
        elif tag.lower() in {"p", "div", "br", "li", "tr", "h1", "h2", "h3", "h4", "h5", "h6"}:
            self.parts.append("\n")

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() in {"script", "style", "noscript", "svg"} and self.suppressed:
            self.suppressed -= 1
        elif tag.lower() in {"p", "div", "li", "tr"}:
            self.parts.append("\n")

    def handle_data(self, data: str) -> None:
        if not self.suppressed:
            self.parts.append(data)

    def text(self) -> str:
        raw = " ".join(self.parts)
        raw = re.sub(r"[\t\f\v ]+", " ", raw)
        raw = re.sub(r"\n\s*\n+", "\n\n", raw)
        return raw.strip() + "\n"


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def json_bytes(value: Any) -> bytes:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")


def iso_date(value: Any, name: str, allow_null: bool = True) -> str | None:
    if value is None and allow_null:
        return None
    if not isinstance(value, str):
        raise IngestError(f"{name} must be an ISO date string")
    try:
        return date.fromisoformat(value).isoformat()
    except ValueError as exc:
        raise IngestError(f"{name} must be YYYY-MM-DD") from exc


def clean_extension(uri: str, content_type: str | None) -> str:
    path = urllib.parse.urlparse(uri).path if urllib.parse.urlparse(uri).scheme else uri
    suffix = Path(path).suffix.lower()
    if suffix and len(suffix) <= 10 and re.fullmatch(r"\.[a-z0-9]+", suffix):
        return suffix
    if content_type:
        guessed = mimetypes.guess_extension(content_type.split(";", 1)[0].strip())
        if guessed:
            return guessed
    return ".bin"


def decode_text(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1254", "windows-1254", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise IngestError("text decoding failed")


def extract_pdf(data: bytes) -> tuple[str | None, str, str | None]:
    try:
        from pypdf import PdfReader  # type: ignore

        reader = PdfReader(io.BytesIO(data))
        pages = [(page.extract_text() or "") for page in reader.pages]
        text = "\n\n".join(pages).strip()
        if text:
            return text + "\n", "extracted", "pypdf"
    except Exception:
        pass

    executable = shutil.which("pdftotext")
    if executable:
        with tempfile.TemporaryDirectory(prefix="muhasebecim-pdf-") as temp_dir:
            source = Path(temp_dir) / "source.pdf"
            target = Path(temp_dir) / "source.txt"
            source.write_bytes(data)
            process = subprocess.run(
                [executable, "-layout", str(source), str(target)],
                capture_output=True,
                text=True,
                timeout=120,
                check=False,
            )
            if process.returncode == 0 and target.exists():
                text = target.read_text(encoding="utf-8", errors="replace").strip()
                if text:
                    return text + "\n", "extracted", "pdftotext"
    return None, "extraction_pending", None


def _spreadsheet_cell_text(value: Any) -> str:
    """Serialize a cell without evaluating formulas or losing visible precision."""
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float):
        return str(value)
    return str(value)


def extract_xlsx(data: bytes) -> tuple[str | None, str, str | None]:
    """Extract XLSX/XLSM cells locally; formulas are preserved and never executed."""
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        return None, "extraction_pending", None

    try:
        workbook = load_workbook(
            io.BytesIO(data),
            read_only=True,
            data_only=False,
            keep_links=False,
        )
        lines: list[str] = []
        for worksheet in workbook.worksheets:
            lines.append(json.dumps({"sheet": worksheet.title}, ensure_ascii=False, sort_keys=True))
            for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                cells = [_spreadsheet_cell_text(value) for value in row]
                while cells and cells[-1] == "":
                    cells.pop()
                if cells:
                    lines.append(
                        json.dumps(
                            {"row": row_number, "cells": cells},
                            ensure_ascii=False,
                            sort_keys=True,
                        )
                    )
        workbook.close()
        text = "\n".join(lines).strip()
        if text:
            return text + "\n", "extracted", "openpyxl-no-formula-evaluation"
        return "", "low_text", "openpyxl-no-formula-evaluation"
    except Exception:
        return None, "extraction_pending", None


def extract_text(data: bytes, extension: str, content_type: str | None) -> tuple[str | None, str, str | None]:
    mime = (content_type or "").lower()
    if extension in {".html", ".htm"} or "text/html" in mime:
        parser = VisibleTextParser()
        parser.feed(decode_text(data))
        return parser.text(), "extracted", "html.parser"
    if extension == ".pdf" or "application/pdf" in mime:
        return extract_pdf(data)
    if extension in {".xlsx", ".xlsm"} or "spreadsheetml" in mime:
        return extract_xlsx(data)
    if extension in TEXT_TYPES or mime.startswith("text/") or "application/json" in mime or "application/xml" in mime:
        return decode_text(data).replace("\r\n", "\n").replace("\r", "\n"), "extracted", "local-decode"
    return None, "unsupported", None


def fetch_document(
    document: dict[str, Any],
    manifest_dir: Path,
    allowed_hosts: set[str],
    download_timeout_seconds: int = 60,
) -> tuple[bytes, dict[str, Any]]:
    uri = document["uri"]
    scope = document["scope"]
    parsed = urllib.parse.urlparse(uri)
    if parsed.scheme:
        if parsed.scheme != "https":
            raise IngestError("only HTTPS URLs are accepted")
        if scope != "public":
            raise IngestError("scope=case documents must be local paths; network upload/fetch is forbidden")
        if parsed.username or parsed.password:
            raise IngestError("URLs containing credentials are forbidden")
        host = (parsed.hostname or "").lower()
        if not host:
            raise IngestError("URL host is missing")
        if allowed_hosts and host not in allowed_hosts and not any(host.endswith("." + item) for item in allowed_hosts):
            raise IngestError(f"host is not in allowed_hosts: {host}")
        request = urllib.request.Request(uri, headers={"User-Agent": f"MuhasebecimLocalIngest/{VERSION}"})
        try:
            with urllib.request.urlopen(request, timeout=download_timeout_seconds) as response:
                data = response.read()
                final_url = response.geturl()
                final_host = (urllib.parse.urlparse(final_url).hostname or "").lower()
                if allowed_hosts and final_host not in allowed_hosts and not any(final_host.endswith("." + item) for item in allowed_hosts):
                    raise IngestError(f"redirected host is not in allowed_hosts: {final_host}")
                return data, {
                    "resolved_uri": final_url,
                    "content_type": response.headers.get("Content-Type"),
                    "etag": response.headers.get("ETag"),
                    "last_modified": response.headers.get("Last-Modified"),
                }
        except (urllib.error.URLError, TimeoutError) as exc:
            raise IngestError(f"download failed for {uri}: {exc}") from exc

    local_path = Path(uri)
    if not local_path.is_absolute():
        local_path = (manifest_dir / local_path).resolve()
    else:
        local_path = local_path.resolve()
    if not local_path.is_file():
        raise IngestError(f"local file does not exist: {local_path}")
    return local_path.read_bytes(), {
        "resolved_uri": str(local_path),
        "content_type": mimetypes.guess_type(local_path.name)[0],
        "etag": None,
        "last_modified": datetime.fromtimestamp(local_path.stat().st_mtime, timezone.utc).isoformat(),
    }


def validate_document(raw: Any, index: int) -> dict[str, Any]:
    if not isinstance(raw, dict):
        raise IngestError(f"documents[{index}] must be an object")
    required = ("uri", "authority", "title", "document_type", "status", "scope")
    for key in required:
        if not isinstance(raw.get(key), str) or not raw[key].strip():
            raise IngestError(f"documents[{index}].{key} is required")
    if raw["status"] not in STATUSES:
        raise IngestError(f"documents[{index}].status must be one of {sorted(STATUSES)}")
    if raw["scope"] not in SCOPES:
        raise IngestError(f"documents[{index}].scope must be one of {sorted(SCOPES)}")
    tags = raw.get("tags", [])
    if not isinstance(tags, list) or not all(isinstance(tag, str) and tag for tag in tags):
        raise IngestError(f"documents[{index}].tags must be a list of strings")
    result = dict(raw)
    result["publication_date"] = iso_date(raw.get("publication_date"), f"documents[{index}].publication_date")
    result["effective_from"] = iso_date(raw.get("effective_from"), f"documents[{index}].effective_from")
    result["effective_to"] = iso_date(raw.get("effective_to"), f"documents[{index}].effective_to")
    result["tags"] = sorted(set(tags))
    return result


def load_manifest(path: Path) -> dict[str, Any]:
    try:
        manifest = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise IngestError(f"cannot read manifest: {exc}") from exc
    if not isinstance(manifest, dict):
        raise IngestError("manifest root must be an object")
    as_of_date = iso_date(manifest.get("as_of_date"), "as_of_date", allow_null=False)
    raw_documents = manifest.get("documents")
    if not isinstance(raw_documents, list) or not raw_documents:
        raise IngestError("documents must be a non-empty list")
    allowed_hosts = manifest.get("allowed_hosts", [])
    if not isinstance(allowed_hosts, list) or not all(isinstance(host, str) and host for host in allowed_hosts):
        raise IngestError("allowed_hosts must be a list of host names")
    download_timeout_seconds = manifest.get("download_timeout_seconds", 60)
    if (
        isinstance(download_timeout_seconds, bool)
        or not isinstance(download_timeout_seconds, int)
        or not 1 <= download_timeout_seconds <= 600
    ):
        raise IngestError("download_timeout_seconds must be an integer between 1 and 600")
    return {
        "as_of_date": as_of_date,
        "allowed_hosts": sorted({host.lower() for host in allowed_hosts}),
        "download_timeout_seconds": download_timeout_seconds,
        "documents": [validate_document(item, index) for index, item in enumerate(raw_documents)],
    }


def load_index(index_path: Path) -> list[dict[str, Any]]:
    if not index_path.exists():
        return []
    records = []
    for line_number, line in enumerate(index_path.read_text(encoding="utf-8").splitlines(), start=1):
        if not line.strip():
            continue
        try:
            record = json.loads(line)
        except json.JSONDecodeError as exc:
            raise IngestError(f"invalid index JSON at line {line_number}") from exc
        records.append(record)
    return records


def ensure_corpus(corpus: Path) -> None:
    for name in ("blobs", "text", "records"):
        (corpus / name).mkdir(parents=True, exist_ok=True)


def ingest_one(
    document: dict[str, Any],
    manifest: dict[str, Any],
    manifest_dir: Path,
    corpus: Path,
    existing: list[dict[str, Any]],
) -> dict[str, Any]:
    data, fetch_meta = fetch_document(
        document,
        manifest_dir,
        set(manifest["allowed_hosts"]),
        manifest["download_timeout_seconds"],
    )
    blob_hash = sha256_bytes(data)
    extension = clean_extension(fetch_meta["resolved_uri"], fetch_meta["content_type"])
    blob_rel = Path("blobs") / f"{blob_hash}{extension}"
    blob_path = corpus / blob_rel
    if not blob_path.exists():
        blob_path.write_bytes(data)

    text, extraction_status, extractor = extract_text(data, extension, fetch_meta["content_type"])
    if text is not None and extraction_status == "extracted" and len(text.strip()) < 200:
        extraction_status = "low_text"
    text_bytes = text.encode("utf-8") if text is not None else None
    text_hash = sha256_bytes(text_bytes) if text_bytes is not None else None
    text_rel = Path("text") / f"{blob_hash}.txt" if text is not None else None
    if text_rel is not None and not (corpus / text_rel).exists():
        (corpus / text_rel).write_bytes(text_bytes or b"")

    identity = {
        "uri": document["uri"],
        "blob_sha256": blob_hash,
        "text_sha256": text_hash,
        "status": document["status"],
        "effective_from": document["effective_from"],
        "effective_to": document["effective_to"],
        "scope": document["scope"],
        "as_of_date": manifest["as_of_date"],
    }
    for old in existing:
        if all(old.get(key) == value for key, value in identity.items()):
            return {"record_id": old["record_id"], "action": "duplicate", "blob_sha256": blob_hash}

    same_uri = [item for item in existing if item.get("uri") == document["uri"]]
    supersedes = same_uri[-1]["record_id"] if same_uri else None
    ingested_at = datetime.now(timezone.utc).isoformat()
    record_seed = {**identity, "ingested_at": ingested_at, "title": document["title"]}
    record_id = hashlib.sha256(json_bytes(record_seed)).hexdigest()[:24]
    record = {
        "record_id": record_id,
        "uri": document["uri"],
        "resolved_uri": fetch_meta["resolved_uri"],
        "authority": document["authority"],
        "title": document["title"],
        "document_type": document["document_type"],
        "publication_date": document["publication_date"],
        "effective_from": document["effective_from"],
        "effective_to": document["effective_to"],
        "status": document["status"],
        "scope": document["scope"],
        "tags": document["tags"],
        "pinpoint_hint": document.get("pinpoint_hint"),
        "as_of_date": manifest["as_of_date"],
        "ingested_at": ingested_at,
        "content_type": fetch_meta["content_type"],
        "etag": fetch_meta["etag"],
        "last_modified": fetch_meta["last_modified"],
        "blob_path": blob_rel.as_posix(),
        "blob_sha256": blob_hash,
        "blob_size": len(data),
        "text_path": text_rel.as_posix() if text_rel else None,
        "text_sha256": text_hash,
        "text_characters": len(text) if text is not None else 0,
        "extraction_status": extraction_status,
        "extractor": extractor,
        "supersedes": supersedes,
    }
    record_path = corpus / "records" / f"{record_id}.json"
    record_path.write_bytes((json.dumps(record, ensure_ascii=False, indent=2) + "\n").encode("utf-8"))
    with (corpus / "index.jsonl").open("a", encoding="utf-8", newline="\n") as handle:
        handle.write(json.dumps(record, ensure_ascii=False, sort_keys=True) + "\n")
    existing.append(record)
    return {"record_id": record_id, "action": "ingested", "blob_sha256": blob_hash, "extraction_status": extraction_status}


def command_ingest(manifest_path: Path, corpus: Path) -> dict[str, Any]:
    manifest_path = manifest_path.resolve()
    corpus = corpus.resolve()
    ensure_corpus(corpus)
    manifest = load_manifest(manifest_path)
    existing = load_index(corpus / "index.jsonl")
    results = []
    errors = []
    for document in manifest["documents"]:
        try:
            results.append(ingest_one(document, manifest, manifest_path.parent, corpus, existing))
        except (IngestError, OSError, subprocess.SubprocessError) as exc:
            errors.append({"uri": document["uri"], "error": str(exc)})
    quality_warnings = [
        {"record_id": item["record_id"], "warning": item["extraction_status"]}
        for item in results
        if item.get("extraction_status") in {"low_text", "extraction_pending", "unsupported"}
    ]
    return {
        "version": VERSION,
        "corpus": str(corpus),
        "ingested": sum(item["action"] == "ingested" for item in results),
        "duplicates": sum(item["action"] == "duplicate" for item in results),
        "results": results,
        "errors": errors,
        "quality_warnings": quality_warnings,
        "ok": not errors,
    }


def command_audit(corpus: Path) -> dict[str, Any]:
    corpus = corpus.resolve()
    records = load_index(corpus / "index.jsonl")
    errors = []
    for record in records:
        blob = corpus / record["blob_path"]
        if not blob.is_file():
            errors.append({"record_id": record["record_id"], "error": "missing blob"})
        elif sha256_bytes(blob.read_bytes()) != record["blob_sha256"]:
            errors.append({"record_id": record["record_id"], "error": "blob hash mismatch"})
        text_path = record.get("text_path")
        if text_path:
            text_file = corpus / text_path
            if not text_file.is_file():
                errors.append({"record_id": record["record_id"], "error": "missing text"})
            elif sha256_bytes(text_file.read_bytes()) != record["text_sha256"]:
                errors.append({"record_id": record["record_id"], "error": "text hash mismatch"})
        record_file = corpus / "records" / f"{record['record_id']}.json"
        if not record_file.is_file():
            errors.append({"record_id": record["record_id"], "error": "missing record file"})
        supersedes = record.get("supersedes")
        if supersedes and not (corpus / "records" / f"{supersedes}.json").is_file():
            errors.append({"record_id": record["record_id"], "error": "missing superseded record"})
    latest_by_uri: dict[str, dict[str, Any]] = {}
    for record in records:
        latest_by_uri[record["uri"]] = record
    latest_status_counts = dict(sorted(Counter(item.get("extraction_status", "unknown") for item in latest_by_uri.values()).items()))
    latest_quality_warnings = [
        {
            "record_id": item["record_id"],
            "uri": item["uri"],
            "warning": item.get("extraction_status", "unknown"),
        }
        for item in latest_by_uri.values()
        if item.get("extraction_status") in {"low_text", "extraction_pending", "unsupported"}
    ]
    latest_as_of_dates = sorted({str(item.get("as_of_date")) for item in latest_by_uri.values()})
    return {
        "version": VERSION,
        "corpus": str(corpus),
        "record_count": len(records),
        "latest_uri_count": len(latest_by_uri),
        "latest_as_of_dates": latest_as_of_dates,
        "latest_extraction_status_counts": latest_status_counts,
        "latest_quality_warnings": latest_quality_warnings,
        "errors": errors,
        "ok": not errors,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    sub = parser.add_subparsers(dest="command", required=True)
    ingest_parser = sub.add_parser("ingest", help="ingest a manifest into a local corpus")
    ingest_parser.add_argument("--manifest", required=True, type=Path)
    ingest_parser.add_argument("--corpus", required=True, type=Path)
    audit_parser = sub.add_parser("audit", help="verify corpus hashes and files")
    audit_parser.add_argument("--corpus", required=True, type=Path)
    args = parser.parse_args()
    try:
        result = command_ingest(args.manifest, args.corpus) if args.command == "ingest" else command_audit(args.corpus)
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return 0 if result["ok"] else 2
    except (IngestError, OSError) as exc:
        print(json.dumps({"ok": False, "error": str(exc)}, ensure_ascii=False), file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
